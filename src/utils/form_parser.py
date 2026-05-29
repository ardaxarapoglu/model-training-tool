"""Parse experiment xlsx forms to extract time-frame PB concentration labels."""
import os
from pathlib import Path

import openpyxl


def parse_experiment_form(xlsx_path: str) -> dict:
    """Return experiment metadata and 7 time-frame records from an xlsx form.

    Kinetik_Numuneler sheet layout (fixed):
        rows 8–14  → Y1…Y7
        col A (1)  → frame name
        col B (2)  → time interval
        col D (4)  → Pb Tenörü (%)  ← label
        col E (5)  → Pb Dağılım (%)
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    result = {
        "experiment_no": "",
        "date": "",
        "operator": "",
        "time_frames": [],
    }

    if "Deney_Bilgileri" in wb.sheetnames:
        ws = wb["Deney_Bilgileri"]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1]
            if key == "Deney No":
                result["experiment_no"] = str(val) if val is not None else ""
            elif key == "Tarih":
                result["date"] = str(val) if val is not None else ""
            elif key in ("Operatör", "Operator"):
                result["operator"] = str(val) if val is not None else ""

    if "Kinetik_Numuneler" in wb.sheetnames:
        ws = wb["Kinetik_Numuneler"]
        # Y1–Y7 are in fixed rows 8–14; Pb Tenörü (%) is column D (index 4 in 1-based, cell D8:D14)
        for excel_row in range(8, 15):  # rows 8, 9, 10, 11, 12, 13, 14
            name     = ws.cell(excel_row, 1).value  # col A
            pb_grade = ws.cell(excel_row, 4).value  # col D  ← Pb Tenörü (%)
            result["time_frames"].append(
                {
                    "name": str(name) if name is not None else f"Y{excel_row - 7}",
                    "pb_concentration": float(pb_grade) if pb_grade is not None else 0.0,
                    "folder_path": "",
                    "notes": "",
                }
            )

    return result


def auto_assign_folders(experiment: dict, frames_root: str) -> dict:
    """Fill folder_path for each time frame by matching Y1…Y7 sub-directories."""
    exp_dir = os.path.join(frames_root, experiment.get("id", ""))
    if not os.path.isdir(exp_dir):
        return experiment
    for tf in experiment.get("time_frames", []):
        candidate = os.path.join(exp_dir, tf["name"])
        if os.path.isdir(candidate):
            tf["folder_path"] = candidate
    return experiment


def discover_experiments(forms_dir: str, frames_dir: str) -> list:
    """Scan forms/ and frames/ directories and return a list of experiment dicts."""
    import uuid

    experiments = []
    if not os.path.isdir(forms_dir):
        return experiments

    for fname in sorted(os.listdir(forms_dir)):
        if not fname.lower().endswith(".xlsx"):
            continue
        exp_id = Path(fname).stem  # e.g., "D01"
        xlsx_path = os.path.join(forms_dir, fname)
        try:
            info = parse_experiment_form(xlsx_path)
        except Exception:
            info = {"experiment_no": exp_id, "date": "", "operator": "", "time_frames": []}

        # Ensure 7 time frames
        existing_names = {tf["name"] for tf in info["time_frames"]}
        for i in range(1, 8):
            yn = f"Y{i}"
            if yn not in existing_names:
                info["time_frames"].append(
                    {
                        "name": yn,
                        "pb_concentration": 0.0,
                        "folder_path": "",
                        "notes": "",
                    }
                )
        info["time_frames"].sort(key=lambda tf: tf["name"])

        # Auto-assign frame folders
        exp_frames_dir = os.path.join(frames_dir, exp_id)
        for tf in info["time_frames"]:
            candidate = os.path.join(exp_frames_dir, tf["name"])
            if os.path.isdir(candidate):
                tf["folder_path"] = candidate

        experiments.append(
            {
                "id": exp_id,
                "name": f"{exp_id} - {info['experiment_no']}",
                "form_path": xlsx_path,
                "experiment_no": info["experiment_no"],
                "date": info["date"],
                "operator": info["operator"],
                "split": "train",
                "notes": "",
                "time_frames": info["time_frames"],
            }
        )

    return experiments
